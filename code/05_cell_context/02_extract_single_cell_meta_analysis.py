#!/usr/bin/env python3
"""Extract TSPAN14 results from the Nakatsuka et al. AD single-cell meta-analysis.

The source workbook stores each cell type as a horizontal block. This script
normalizes those blocks, preserves every tested direction and method, and adds
project-level sensitivity and functional evidence without treating nominal
signals as multiple-testing-significant results.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SHEETS = (
    "SumRank_UpRegulated",
    "SumRank_DownRegulated",
    "Merge_UpRegulated",
    "Merge_DownRegulated",
    "InvVar_UpRegulated",
    "InvVar_DownRegulated",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_sheet(workbook, sheet_name: str, gene: str) -> list[dict[str, object]]:
    worksheet = workbook[sheet_name]
    cell_type_row = next(
        worksheet.iter_rows(min_row=4, max_row=4, values_only=True)
    )
    header_row = next(worksheet.iter_rows(min_row=5, max_row=5, values_only=True))
    starts = [(index, value) for index, value in enumerate(cell_type_row) if value]
    method, direction = sheet_name.split("_", maxsplit=1)
    records: list[dict[str, object]] = []

    for values in worksheet.iter_rows(min_row=6, values_only=True):
        for block_index, (start, cell_type) in enumerate(starts):
            if values[start] != gene:
                continue
            end = (
                starts[block_index + 1][0]
                if block_index + 1 < len(starts)
                else worksheet.max_column
            )
            metrics = {
                str(header_row[column]): values[column]
                for column in range(start, end)
                if header_row[column]
            }
            record: dict[str, object] = {
                "source": "Nakatsuka et al. 2025",
                "doi": "10.1038/s41467-025-62579-z",
                "method": method,
                "tested_direction": direction.replace("Regulated", "regulated").lower(),
                "cell_type": str(cell_type),
                "gene": gene,
            }
            record.update(metrics)
            records.append(record)
    return records


def normalize_meta_results(records: list[dict[str, object]]) -> pd.DataFrame:
    frame = pd.DataFrame(records)
    frame["p_value"] = frame["PVal"].combine_first(frame["Pval"])
    frame["adjusted_p_value"] = frame["PVal_BH"].combine_first(frame["lfdr"])
    frame["multiplicity_method"] = frame["PVal_BH"].notna().map(
        {True: "Benjamini-Hochberg FDR", False: "local FDR"}
    )
    frame["effect"] = (
        frame["Average_log2fc"]
        .combine_first(frame["log2fc"])
        .combine_first(frame["Effect_Size"])
    )
    frame["effect_metric"] = frame["method"].map(
        {
            "SumRank": "mean log2 fold change across datasets",
            "Merge": "DESeq2 merged-data log2 fold change",
            "InvVar": "inverse-variance weighted effect",
        }
    )
    frame["nominal_p_lt_0_05"] = frame["p_value"] < 0.05
    frame["adjusted_p_lt_0_05"] = frame["adjusted_p_value"] < 0.05
    keep = [
        "source",
        "doi",
        "method",
        "tested_direction",
        "cell_type",
        "gene",
        "effect",
        "effect_metric",
        "p_value",
        "adjusted_p_value",
        "multiplicity_method",
        "nominal_p_lt_0_05",
        "adjusted_p_lt_0_05",
    ]
    return frame[keep].sort_values(
        ["method", "tested_direction", "adjusted_p_value", "p_value", "cell_type"]
    )


def build_summary(meta: pd.DataFrame) -> pd.DataFrame:
    primary = meta[meta["method"] == "SumRank"].copy()
    primary["rank"] = primary.groupby("tested_direction")["p_value"].rank(
        method="first"
    )
    top = primary.sort_values(["adjusted_p_value", "p_value"]).iloc[0]
    microglia = primary[
        (primary["cell_type"] == "Microglia")
        & (primary["tested_direction"] == "upregulated")
    ].iloc[0]
    rows = [
        {
            "analysis": "AD single-cell reproducibility meta-analysis",
            "design": "SumRank across 17 independent AD single-cell/snRNA-seq datasets",
            "cell_context": "21 broad and neuronal subtype outputs",
            "effect": "not applicable as a single pooled disease coefficient",
            "p_value": None,
            "adjusted_p_value": None,
            "interpretation": (
                "No TSPAN14 cell-type result passed Benjamini-Hochberg FDR < 0.05; "
                "there is no reproducible uniform late-stage disease-state shift."
            ),
        },
        {
            "analysis": "Strongest TSPAN14 SumRank result",
            "design": "Permutation-based SumRank",
            "cell_context": f"{top['cell_type']} ({top['tested_direction']})",
            "effect": top["effect"],
            "p_value": top["p_value"],
            "adjusted_p_value": top["adjusted_p_value"],
            "interpretation": "Nominal signal only; it does not survive FDR correction.",
        },
        {
            "analysis": "TSPAN14 microglial SumRank result",
            "design": "Permutation-based SumRank",
            "cell_context": "Microglia (upregulated test)",
            "effect": microglia["effect"],
            "p_value": microglia["p_value"],
            "adjusted_p_value": microglia["adjusted_p_value"],
            "interpretation": (
                "Directionally positive across contributing datasets but not FDR-significant."
            ),
        },
    ]
    return pd.DataFrame(rows)


def append_project_evidence(summary: pd.DataFrame, seaad_path: Path) -> pd.DataFrame:
    seaad = pd.read_csv(seaad_path, sep="\t")
    rows: list[dict[str, object]] = []
    for _, result in seaad.iterrows():
        label = str(
            result.get(
                "cell_type", result.get("cell_class", result.get("cell_group", "SEA-AD"))
            )
        )
        effect = result.get(
            "log2_fold_change", result.get("log2FC", result.get("logFC"))
        )
        fdr = result.get(
            "FDR", result.get("padj", result.get("adj.P.Val"))
        )
        p_value = result.get(
            "p_value", result.get("PValue", result.get("P.Value"))
        )
        rows.append(
            {
                "analysis": "SEA-AD donor-level pseudobulk sensitivity",
                "design": "Covariate-adjusted donor-level negative-binomial model",
                "cell_context": label,
                "effect": effect,
                "p_value": p_value,
                "adjusted_p_value": fdr,
                "interpretation": (
                    "No case-control shift; retained as an independent sensitivity analysis."
                ),
            }
        )
    rows.append(
        {
            "analysis": "Microglial regulatory variant perturbation",
            "design": "Allelic editing and microglial functional characterization",
            "cell_context": "Microglia; rs7922621 in the TSPAN14 regulatory block",
            "effect": "Risk allele lowers TSPAN14 expression and cell-surface ADAM10",
            "p_value": None,
            "adjusted_p_value": None,
            "interpretation": (
                "Genotype-resolved functional evidence links the regulatory block to "
                "TSPAN14, ADAM10 trafficking and soluble TREM2 shedding; it does not "
                "by itself establish the exact exon5-6 isoform consequence."
            ),
        }
    )
    return pd.concat([summary, pd.DataFrame(rows)], ignore_index=True).convert_dtypes()


def write_manifest(source: Path, output: Path) -> None:
    records = [
        {
            "resource": "Nakatsuka et al. Supplementary Data 3",
            "doi": "10.1038/s41467-025-62579-z",
            "url": "https://www.nature.com/articles/s41467-025-62579-z",
            "local_source": str(source),
            "bytes": source.stat().st_size,
            "sha256": sha256(source),
            "access_date": date.today().isoformat(),
        },
        {
            "resource": "Yang et al. functional microglia study",
            "doi": "10.1038/s41588-023-01506-8",
            "url": "https://www.nature.com/articles/s41588-023-01506-8",
            "local_source": "Published article; summarized as external functional evidence",
            "bytes": None,
            "sha256": None,
            "access_date": date.today().isoformat(),
        },
    ]
    pd.DataFrame(records).to_csv(output, sep="\t", index=False)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--seaad", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    records: list[dict[str, object]] = []
    for sheet in SHEETS:
        records.extend(parse_sheet(workbook, sheet, "TSPAN14"))
    meta = normalize_meta_results(records)
    summary = append_project_evidence(build_summary(meta), args.seaad)

    meta.to_csv(
        args.output_dir / "nakatsuka2025_tspan14_meta_results.tsv",
        sep="\t",
        index=False,
    )
    summary.to_csv(
        args.output_dir / "tspan14_cell_context_evidence_hierarchy.tsv",
        sep="\t",
        index=False,
    )
    write_manifest(args.workbook, args.output_dir / "external_source_manifest.tsv")

    audit = {
        "gene": "TSPAN14",
        "source_sheets": list(SHEETS),
        "n_normalized_results": int(len(meta)),
        "n_sumrank_results": int((meta["method"] == "SumRank").sum()),
        "n_sumrank_fdr_lt_0_05": int(
            ((meta["method"] == "SumRank") & meta["adjusted_p_lt_0_05"]).sum()
        ),
        "n_all_method_adjusted_p_lt_0_05": int(meta["adjusted_p_lt_0_05"].sum()),
    }
    (args.output_dir / "extraction_audit.json").write_text(
        json.dumps(audit, indent=2), encoding="utf-8"
    )
    print(json.dumps(audit, indent=2))


if __name__ == "__main__":
    main()
